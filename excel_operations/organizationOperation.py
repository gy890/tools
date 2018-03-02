# coding=utf-8
"""
Created on 2016年12月15日

@author: EP-Admin
"""
import os
import time
from openpyxl import load_workbook


def timeit(func):
    def _deco(*args, **kwargs):
        t1 = time.time()
        ret = func(*args, **kwargs)
        t2 = time.time()
        #         logger.debug('{0} consume time: {1:.3f} s'.format(func.__name__, (t2 - t1)))
        print('{0} consume time: {1:.3f} s'.format(func.__name__, (t2 - t1)))
        return ret

    return _deco


def get_invalid(xlsx):
    wb = load_workbook(xlsx)
    sheet_names = wb.sheetnames[1:]
    ws = wb.get_sheet_by_name(sheet_names[0])
    if ws.cell(row=1, column=2).value == '公司英文名称':
        return os.path.split(xlsx)[1]
    else:
        return -1


def get_organizations(xlsx):
    wb = load_workbook(xlsx)
    sheet_names = wb.sheetnames[1:]
    organizations_names = []
    for each in sheet_names:
        ws = wb.get_sheet_by_name(each)
        organization_name = ws.cell(row=2, column=1).value
        if organization_name is not None:
            organizations_names.append(organization_name)
        else:
            print(os.path.split(xlsx)[1], each)
            # 没公司名字把sheet名当作名字，save表格比较耗时间
            ws.cell(row=2, column=1).value = each
            wb.save(xlsx)
            organizations_names.append(each)
    return os.path.split(xlsx)[1], organizations_names


@timeit
def main():
    """检测公司名，写到organizations.txt中

    """
    path = input("Enter the path: ")
    with open('organizations.txt', 'w', encoding='utf-8') as f:
        total = 0
        for i, each in enumerate(os.listdir(path), 1):
            print("{} in process...".format(i))
            if each.endswith('.xlsx'):
                filename = os.path.join(path, each)
                (city, organizations) = get_organizations(filename)
                total += int(len(organizations))
                for o in organizations:
                    if o is None:
                        print(filename)
                    f.write(city + '\t' + o.strip() + '\n')
        print(total)


@timeit
def main2():
    """检测模板用错的表格，写到invalid.txt中

    """
    path = input("Enter the path: ")
    with open('invalid.txt', 'w', encoding='utf-8') as f:
        total = 0
        for i, each in enumerate(os.listdir(path), 1):
            print("{} in process...".format(i))
            if each.endswith('.xlsx'):
                filename = os.path.join(path, each)
                valid = get_invalid(filename)
                if valid != -1:
                    f.write(valid + '\n')
                else:
                    total += 1
        print(total)


if __name__ == '__main__':
    main2()
